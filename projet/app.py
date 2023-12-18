#Tous les biblio nécessaire
import math
import cv2
from flask import Flask, Response,session,abort,render_template,redirect,request,jsonify, url_for
import os,pathlib
#app.py
import os
import pathlib
import requests
from pip._vendor import cachecontrol
import openpyxl as xl;
import pyrebase
import firebase_admin
from firebase_admin import auth
from firebase_admin import storage,credentials
import pandas as pd
import jpype
import asposecells
from datetime import datetime
from firebase_admin import db
import json
from firebase_admin import credentials, messaging
import win32com
jpype.startJVM()
from firebase_admin import firestore
from asposecells.api import Workbook
from flask_cors import CORS
from flask_mail import Mail
from flask_mail import Message
import secrets
import string
from openpyxl.styles.colors import Color
from openpyxl import load_workbook
from flask import Flask, request
from flask import Flask
from openpyxl.styles import PatternFill

import pytesseract
from PIL import Image
import pandas as pd
from PIL import Image
app = Flask(__name__)

app=Flask("GMPF")
CORS(app)

#lieé l'application avec firebase
config={
  "apiKey": "AIzaSyAOrH7iSoSR5qc5IF7cpPwkhY4LSNKZvIg",
  "authDomain": "gmpf-68019.firebaseapp.com",
  "projectId": "gmpf-68019",
  "storageBucket": "gmpf-68019.appspot.com",
  "messagingSenderId": "814290862383",
  "appId": "1:814290862383:web:64d51511afee99a946a1da",
  "measurementId": "G-MX09RW75P0",
  "databaseURL":"https://gmpf-68019-default-rtdb.firebaseio.com"
}
# Initialiser le SDK d'administration Firebase
cred = credentials.Certificate('projet/servicesAccountKey.json')
firebase_admin.initialize_app(cred, {
        'storageBucket': 'gmpf-68019.appspot.com',
        "databaseURL":"https://gmpf-68019-default-rtdb.firebaseio.com"
})
#send mail config
app.config['MAIL_SERVER'] = 'smtp.gmail.com'
app.config['MAIL_PORT'] = 587
app.config['MAIL_USE_TLS'] = True
app.config['MAIL_USERNAME'] = 'fares.elfehri10@gmail.com'
app.config['MAIL_PASSWORD'] = 'zbqapesvatjstbcw'
mail = Mail(app)
# Obtenir une référence au bucket Firebase Storage
bucket = storage.bucket()
#ici nous faisons l'authentification firebase
firebase=pyrebase.initialize_app(config)
database=firebase.database()
db = firestore.client()
Auth=firebase.auth()

from google.oauth2 import service_account
from googleapiclient.discovery import build
from googleapiclient.errors import HttpError
import os

@app.route('/upload', methods=['POST'])
def upload_file():
    file = request.files['file']
    temp_file_path = f"{file.filename}"
    file.save(temp_file_path)
    return 'File uploaded successfully'


app.secret_key = 'faresA123'







@app.route('/login',methods = ['POST'])
def login():
    if request.method =='POST' :
              data = request.data
              if request.form.get('login') == 'login': 
                  email=request.form.get('email')
                  password=request.form.get('password')
                  user = Auth.sign_in_with_email_and_password(email, password)
              elif request.form.get('register') == 'register':
                  create_user()
    return jsonify({'reponse':'connecté'})
#Creer un utilisateur
def create_user(name,email,password):
    if request.method =='POST' :
        userdata = {
                'display_name':name,
                'email':email,
                'password':password,
            }
        try:
            user = auth.get_user(email)
        except:   
            user = auth.create_user(**userdata)
            user = auth.get_user_by_email(email)
            doc_ref = db.collection('users').document("list").collection('partenaire').document(user.uid)
            doc_ref.set({name:userdata})
    return jsonify({'reponse ': 'success'})
#Creer un admin
@app.route('/register_admin',methods = ['POST'])
def create_admin():
    if request.method =='POST' :
        admin_user = auth.create_user(
            email=request.form.get('email'),
            password=request.form.get('password'),
            display_name=request.form.get('display_name'),
            email_verified=True,
            disabled=False
        )
        auth.set_custom_user_claims(admin_user.uid, {'admin': True})
    return jsonify({'reponse':'admin est creer'})
@app.route('/users/<user_id>', methods=['GET'])
def get_user_info(user_id):
        user = auth.get_user(user_id)
        user_info = {
            "uid": user.uid,
            "email": user.email,
            "email_verified": user.email_verified,
            "display_name": user.display_name,
            "disabled": user.disabled,
        }
        return jsonify(user_info)

@app.route('/listusers',methods=["GET"])
def get_list_users():
    users = auth.list_users()
    user_list = []
    for user in users.users:
        user_dict = {
            'email': user.email,
            'display_name': user.display_name
        }
        user_list.append(user_dict)
    return jsonify({'users': user_list})
#Ajouter un fichier sur firebase (ADMIN)
def add_file_in_firebase_admin(file_path,name)  :
  # Upload a file to the bucket
    blob = bucket.blob(f'{"admin"}/{name}')
    blob.upload_from_filename(file_path)
    return 'File uploaded to Firebase Storage'
def implement_file(file_path,name):
    # Upload a file to the bucket
    blob = bucket.blob(f'{"admin"}/{name}.xlsx')
    blob.upload_from_file(file_path)
    return 'fichier racine est bien charger dans firebase'
#Creer automatiquement le fichier 
def create_file(name):
    w = Workbook()
    # Save the workbook
    w.save(f'{name}.xlsx')
    # Upload the file to Firebase Storage
    bucket = storage.bucket()
    blob = bucket.blob(f'{"users"}/{name}/{name}.xlsx')
    blob.upload_from_filename(filename=f'{name}.xlsx')
    return name
#Ajouter un fichier sur firebase (USER)
def add_file_in_firebase_user(file_path,name,sheet,dossier):    
  # Upload a file to the bucket
    blob1 = bucket.blob(f'{"admin/"+dossier}/{sheet}.xlsx')
    blob = bucket.blob(f'{"users"}/{name}/{sheet}.xlsx')
    blob1.upload_from_filename(file_path)
    print('File user uploaded to Firebase Storage')

    blob.upload_from_filename(file_path)
    print('File user uploaded to Firebase Storage')
# l'insertion de feuille dans le ficher racine apres la modification
import xlwings as xw
  
@app.route('/sauvegarder/<string:name>/<string:excelname>/<string:id>/<string:dossier>',methods=['POST','GET'])

def return_sheet(name,excelname,id,dossier):
    add_file_in_firebase_user(f"C:/Users/21628/Desktop/test/{id}.xlsx",name,id,dossier)
  

    return {'hi':id}
    
#Sauvegarder la modification 
@app.route('/<string:excelname>/<string:name>',methods=['GET'])
def modifier(excelname,name):
    path=getPath(excelname)
    add_file_in_firebase_user(rf"C:\Users\21628\Downloads\users_{name}_{name}.xlsx",name)
    return {'hi':200}
#Assigner une ou plusieurs feuille a un utilisateur
@app.route('/assigner/<string:name>/<string:excelname>/<string:id>',methods=['POST','GET'])
def assign_sheet(name,excelname,id):    
    stringList = request.args.get('strings')
    myStrings = stringList.split(",")
    index=0
    for sheet in myStrings:
        path=getPath(excelname)
        #destination
        blob = bucket.blob(f'{"users"}/{name}/{name}.xlsx')
        blob.download_to_filename(f'{path}\\{sheet}.xlsx')
        workbook1 = Workbook(f'{path}\{sheet}.xlsx')
        #source
        blob = bucket.blob(fr'admin/{excelname}')
        blob.download_to_filename(f'{path}/fichier racine v0.xlsx')
        workbook2 = Workbook(f'{path}/fichier racine v0.xlsx')
        workbook1.getWorksheets().get(0).copy(workbook2.getWorksheets().get(sheet))
        workbook1.save(f'{path}/{sheet}.xlsx')
        #source
        index+=1
        add_file_in_firebase_user(f'{path}\{sheet}.xlsx',name,sheet,"shared")
    return jsonify({"response":"feuille bien assigner"})
@app.route('/getcoordinate/<string:excelname>',methods=['GET'])
def getCells(excelname):
    path=getPath(excelname)
   # load the Excel workbook
    workbook = xl.load_workbook(path)
    listCells=[]
    # select the worksheet
    worksheet = workbook['Sheet2']
    for row in worksheet.iter_rows(min_row=1, max_row=10 ,min_col=1,max_col=16):
        for cell in row:
            # get the cell name
            cell_name = cell.coordinate
            listCells.append(cell_name)
    result = split_list(listCells, 16)
    return jsonify(result)
def split_list(lst, chunk_size):
    return [lst[i:i+chunk_size] for i in range(0, len(lst), chunk_size)]
import win32com.client as win32
import pythoncom
from openpyxl.styles import Protection
@app.route('/lock/<string:excelname>/<string:id>',methods=['POST','GET'])

def lock(excelname):
    path=getPath(excelname)
    workbook = xl.load_workbook(path+'\\file.xlsx',data_only=True)   
    sheet = workbook.active
    worksheet = workbook['Sheet2']

    cell = sheet['A1']
    cell.protection = Protection(locked=True)
    workbook.save(path+'\\file.xlsx')


def open_excel_file(filename):
    try:
        pythoncom.CoInitialize()  # Initialize the COM library
        excel_app = win32.gencache.EnsureDispatch('Excel.Application')
        workbook = excel_app.Workbooks.Open(filename)
        excel_app.Visible = True  # Show the Excel application window
        # Optional: You can perform further operations on the workbook or its sheets here
        # For example: workbook.Sheets(1).Cells(1, 1).Value = "Hello, Excel!"
    except Exception as e:
        print("An error occurred:", str(e))
    finally:
        excel_app = None  # Release the Excel application object

@app.route('/open/<string:name>/<string:id>',methods=['GET'])
def open(name,id):
    filename = f"C:/Users/21628/Desktop/test/{id}.xlsx"
    open_excel_file(filename)
    return jsonify({"result":200})
@app.route('/getSheets/<string:excelname>',methods=['GET'])
def getAllSheets(excelname):
    # Load the Excel file
    path=getPath(excelname)
    workbook = xl.load_workbook(path+"\\"+"file.xlsx")
    # Get the sheet names
    sheet_names = workbook.sheetnames
    # Print the sheet names
    return jsonify(sheet_names)
@app.route('/maxRowsCols/<string:name>/<string:excelname>/<string:id>',methods=['GET'])
def maxRowsCols(excelname,id,name):
     path=fr"C:\Users\21628\Desktop\test\{id}.xlsx"
     blob = bucket.blob(f"users/{name}/"+excelname)
     blob.download_to_filename(path)   
     workbook = xl.load_workbook(path,data_only=True)
     sheet = workbook["Sheet1"]
     inf={}
     for col in range(1, sheet.max_column + 1):
        cell = sheet.cell(row=1, column=col)
        if cell.value is not None:
            inf["max_cols"]= col
     for r in range(1, sheet.max_row + 1):
        cell = sheet.cell(row=r, column=1)
        if cell.value is not None:  
            inf["max_rows"]= r
        else :
            inf["max_rows"]= sheet.max_row

     print(inf)
     return jsonify(inf)
@app.route('/getDash/<string:excelname>/<string:id>',methods=['POST','GET'])
def getDash(excelname,id):

    path=getPath(excelname)
    workbook = xl.load_workbook(path+'\\file.xlsx',data_only=True)
    #sheet name,path on doit reccuperer a partir de fromulaire admin 
    sheet = workbook[id]  # Replace 'Sheet1' with the name of your sheet

    cell_value = sheet['M35'].value  # Replace 'A1' with the cell reference you want to retrieve

    print(cell_value)
    return jsonify({"":""})
@app.route('/row/<string:excelname>/<string:id>/<int:number>',methods=['POST','GET'])
def selectRow(excelname,id,number):
    path=getPath(excelname)
   # load the Excel workbook
    workbook = xl.load_workbook(path+'\\file.xlsx',data_only=True)
    listCells=[]
    # select the worksheet
    worksheet = workbook[id]
    for row in worksheet.iter_rows(min_row=number, max_row=number ,min_col=worksheet.min_column,max_col=worksheet.max_column):
        for cell in row:
          if cell.value is not None:  
            cell_name = cell.value
            listCells.append(cell_name)
    return jsonify(listCells)

import re

pytesseract.pytesseract.tesseract_cmd = r'C:\Program Files\Tesseract-OCR\tesseract.exe'
def extract_text_from_image(image_path):
    image = Image.open(image_path)
    extracted_text = pytesseract.image_to_string(image)
    return extracted_text
def preprocess_text(text):
    # Remove special characters and unwanted whitespace
    cleaned_text = re.sub(r'[^\w\s]', '', text)
    cleaned_text = re.sub(r'\s+', ' ', cleaned_text).strip()
    return cleaned_text
def extract_table_data(text):
    # Define patterns for table structure
    row_pattern = re.compile(r'\n')
    column_pattern = re.compile(r'\t|\s{2,}')

    # Split the text into rows
    rows = row_pattern.split(text)
    table_data = []

    for row in rows:
        # Split each row into columns
        columns = column_pattern.split(row)
        table_data.append(columns)

    return table_data

def postprocess_table_data(table_data):
    # Remove header row if necessary
    header_row = table_data[0]
    # ...

    # Convert data types or handle empty cells
    # ...

    return table_data
@app.route('/ocr-excel',methods=['POST','GET'])

def create_excel_from_image():
    # Perform OCR on the image
    extracted_text = pytesseract.image_to_string(Image.open(request.form.get('file')))

    # Split the extracted text into rows
    rows = extracted_text.strip().split('\n')

    # Extract headers and data from the rows
    headers = rows[0].split('\t')
    data = [row.split('\t') for row in rows[1:]]

    # Create a DataFrame from the data
    df = pd.DataFrame(data, columns=headers)

    # Save the DataFrame to an Excel file
    df.to_excel(request.form.get('excel'), index=False)
@app.route('/ocr',methods=['POST','GET'])

def extract_table_data_from_image():
    
    # Step 1: Extract text from the image
    extracted_text = extract_text_from_image(request.form.get('file'))

    # Step 2: Preprocess the extracted text
    cleaned_text = preprocess_text(extracted_text)

    # Step 3: Identify the table structure
    table_data = extract_table_data(cleaned_text)

    # Step 4: Post-process the table data
    processed_table_data = postprocess_table_data(table_data)

    return processed_table_data
def preprocess_image(image_path):
    # Load the image using PIL
    image = Image.open(image_path)
    
    # Convert the image to grayscale
    grayscale_image = image.convert("L")
    
    # Apply thresholding to enhance text
    threshold_image = grayscale_image.point(lambda p: p > 150 and 255)
    
    return threshold_image
@app.route('/col/<string:excelname>/<string:id>/<string:col>',methods=['POST','GET'])
def selectCol(excelname,id,col):
    path=getPath(excelname)
    workbook = xl.load_workbook(path+'\\file.xlsx',data_only=True)
   # Select the desired worksheet
    worksheet = workbook[id]  # Replace 'Sheet1' with the name of your sheet
    # Retrieve the value of a specific cell
    cell_value = worksheet[col].value  # Replace 'A1' with the cell you want to retrieve
    return jsonify({'value':cell_value})  # Convert the value to a string and return it as the response
api_endpoint = 'https://api.tabula.technology/tabula'
@app.route('/addsheet',methods=['POST','GET'])
def addexcel():
    data = request.json
    blob = bucket.blob("users/"+data['name']+'/'+data['sheet']+".xlsx")
    blob.download_to_filename(fr"C:\Users\21628\Desktop\test\{data['sheet']}.xlsx")
    path=fr"C:\Users\21628\Desktop\test\{data['sheet']}.xlsx"
    workbook = xl.load_workbook(path,data_only=True)
    print(data)
   # Select the desired worksheet
    worksheet = workbook["Sheet1"]  # Replace 'Sheet1' with the name of your sheet
    from_row_number =int(data['from_row'])
    to_row_number=int(data['to_row'])
    from_col=int(data['from_col'])
    k=from_row_number
    print(from_row_number,from_col,to_row_number)
    while(k<to_row_number):
        for items in data['ocr']:
            j=from_col
            for V in items:
                worksheet.cell(row=k, column=j).value =V
                j+=1
            k+=1
    # Save the modified workbook
    workbook.save(path)
   
    add_file_in_firebase_user(f"C:/Users/21628/Desktop/test/{data['sheet']}.xlsx",data['name'],data['sheet'],"latest")
    # Save the workbook
    workbook.save(path)
    return jsonify({"200":44})

from pytesseract import Output

@app.route('/ocrdict',methods=['POST','GET'])
def up():
    if 'file' not in request.files:
        return 'No file found', 400
    file = request.files['file']
  
    filename = file.filename

    file.save('projet/'+filename)   
    processed_image = preprocess_image('projet/'+filename)
    result = pytesseract.image_to_string(processed_image, output_type=Output.DICT)

    # Chargement de l'image avec openCV
    image = cv2.imread('projet/'+filename)
    # Extract the recognized text
    text = result["text"] 
    img = cv2.imread('projet/'+filename)
    custom_config = r'--oem 3 --psm 6 outputbase digits'

    d = pytesseract.image_to_data(img, output_type=Output.DICT)
    keys = list(d.keys())
    # Print the extracted text
    lines = text.split('\n')
    n_boxes = len(d['text'])
    for i in range(n_boxes):
        if int(d['conf'][i]) > 60:
            (x, y, w, h) = (d['left'][i], d['top'][i], d['width'][i], d['height'][i])
            img = cv2.rectangle(img, (x, y), (x + w, y + h), (0, 255, 0), 2)

    cv2.imshow('img', img)
    cv2.waitKey(0)

    # Remove empty lines from the list
    lines = [line for line in lines if line.strip()] 
    return jsonify(lines)

@app.route('/cell/<string:excelname>/<string:id>/<string:col>',methods=['POST','GET'])
def selectCell(excelname,id,col):
    path=getPath(excelname)
    workbook = xl.load_workbook(path+'\\file.xlsx',data_only=True)
    df = pd.read_excel(path+'\\file.xlsx')
    column = df[col]  # Replace 'Column_Name' with the actual name of the column you want to select
    column_values = column.dropna().tolist()
    return jsonify(column_values)
def getPath(excelname):
    blob = bucket.blob("admin/"+excelname)
    blob.download_to_filename(r"C:\Users\21628\Desktop\test\file.xlsx")    # Load the workbook
    return r"C:\Users\21628\Desktop\test"
@app.route('/getInfo/<string:excelname>',methods=['POST','GET'])
def getFileInf(excelname):
   # Specify the path to the Excel file
    path=getPath(excelname)
    wb = xl.load_workbook(path+'\\file.xlsx',data_only=True)
    # Get the size of the Excel file in bytes
    filesize = os.path.getsize(path)
    # Get the number of sheets in the workbook
    num_sheets = len(wb.sheetnames)
    filename = os.path.basename(path)
    file={}
    # Print the file size and number of sheets
    print(f"File size: {filesize} bytes")
    print(f"Number of sheets: {num_sheets}")
    file['size']=filesize
    file['name']=excelname
    file['Type']='Modele Financier'
    file['Number of sheetes']=num_sheets
    # Print the dashboard data
    return jsonify(file)
@app.route('/getData/<string:name>/<string:excelname>/<string:id>',methods=['POST','GET'])
def getData(excelname,id,name):    
    path=fr"C:\Users\21628\Desktop\test\{id}.xlsx"
    blob = bucket.blob(f"users/{name}/"+excelname)
    blob.download_to_filename(path)   
    workbook = xl.load_workbook(path,data_only=True)
    #sheet name,path on doit reccuperer a partir de fromulaire admin 
    worksheet = workbook["Sheet1"]
    non_null_rows = {}
    i=0
    #min,max rows reccuperer par les chekbox de sheet front 
    for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row,min_col=1, max_col=worksheet.max_column, values_only=True):
        if any(cell is not None for cell in row): 
            row = [cell for cell in row if cell is not None]  # Select non-null values in the row
            non_null_rows[i]=row
            i+=1
    # Print the row values
    return jsonify(non_null_rows)






@app.route('/get/<string:excelname>',methods = ['POST','GET'])
def getInfo(excelname):
    path=getPath(excelname)
    workbook = xl.load_workbook(path+'\\file.xlsx',data_only=True)
    # Load the Excel file
    sheet = workbook['Hyp']
    # Get the values in column A
    column = sheet['M34']
    # Print the values
    print(column.value)

    column = sheet['M35']
    # Print the values
    print(column.value)


@app.route('/sendInvit',methods = ['POST','GET'])
def send_email():
    with app.app_context():
        data = request.get_json()
        password=generatePassword()
        print(data['address'])
        msg = Message("invitation", sender='fares.elfehri10@gmail.com', recipients=[data['address']])
        msg.body = "Salut "+data['name']+" invitation pour rejoindre la plateforme de gestion de pilotages financiers \n lien :http://127.0.0.1:4200/auth/login \n username : "+data['address']+'\n password : '+password
        mail.send(msg)
        create_user(data['name'],data['address'],password)
        create_file(data['name'])
    return {"response":"invitation done"}   

def generatePassword():
     alphabet = string.ascii_letters + string.digits + string.punctuation
     password = ''.join(secrets.choice(alphabet) for i in range(12))
     return password

if __name__=="__main__":
     app.run()